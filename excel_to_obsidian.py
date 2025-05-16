import openpyxl
import os
import re

def create_obsidian_note(question, answer, clarification, category, output_dir="Java Interview Prep"):
    """Создает отдельную заметку Obsidian для вопроса и ответа с уточнением."""
    os.makedirs(output_dir, exist_ok=True)
    # Очищаем имя файла от недопустимых символов
    filename_base = re.sub(r'[^\w\s-]', '', question).strip().replace(' ', '_')
    filename_base = filename_base.replace('\n', '').replace('?', '').replace('(', '').replace(')', '')
    filename = f"{output_dir}/{category}/{filename_base}.md"
    os.makedirs(os.path.dirname(filename), exist_ok=True)
    with open(filename, 'w', encoding='utf-8') as f:
        f.write(f"# {question}\n\n{answer}\n\n**Уточнение:** {clarification}")

def update_index_note(question, category, index_file="Java Interview Map.md"):
    """Обновляет карту знаний Obsidian, добавляя ссылку на новую заметку."""
    filename_base = re.sub(r'[^\w\s-]', '', question).strip().replace(' ', '_')
    filename_base = filename_base.replace('\n', '').replace('?', '').replace('(', '').replace(')', '')
    link = f"[[Java Interview Prep/{category}/{filename_base}]]"
    with open(index_file, 'a', encoding='utf-8') as f:
        f.write(f"* {category}: {link}\n")

def process_excel(excel_filepath):
    """Обрабатывает Excel-файл со столбцами: A - вопрос, B - ответ, C - уточнение."""
    try:
        workbook = openpyxl.load_workbook(excel_filepath)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            category = sheet_name  # Используем название листа как категорию
            for row in sheet.iter_rows(min_row=1, values_only=True):
                if len(row) >= 3 and row[0] and row[1]:  # Проверяем наличие вопроса и ответа
                    question = str(row[0]).strip()
                    answer = str(row[1]).strip()
                    clarification = str(row[2]).strip() if row[2] else ""
                    create_obsidian_note(question, answer, clarification, category)
                    update_index_note(question, category)
                elif len(row) >= 2 and row[0] and row[1]: # Если нет столбца C
                    question = str(row[0]).strip()
                    answer = str(row[1]).strip()
                    clarification = ""
                    create_obsidian_note(question, answer, clarification, category)
                    update_index_note(question, category)
                elif row and row[0]:
                    print(f"Предупреждение: На листе '{category}' отсутствует ответ для вопроса: '{row[0]}'.")

    except FileNotFoundError:
        print(f"Ошибка: Файл '{excel_filepath}' не найден.")
    except Exception as e:
        print(f"Произошла ошибка при обработке Excel-файла: {e}")

if __name__ == "__main__":
    excel_file = "C:\\Users\\ещзы246\\Downloads\\your_interview_questions.xlsx"  # Укажи правильный путь к своему файлу
    process_excel(excel_file)
    print("Заметки Obsidian успешно созданы и карта знаний обновлена!")